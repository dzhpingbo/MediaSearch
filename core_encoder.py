# ── 必须在所有 import 之前禁用代理，否则 httpx 会缓存代理设置 ──
import os
os.environ["http_proxy"] = ""
os.environ["https_proxy"] = ""
os.environ["HTTP_PROXY"] = ""
os.environ["HTTPS_PROXY"] = ""
os.environ["no_proxy"] = "*"
os.environ["NO_PROXY"] = "*"

"""
core_encoder.py - 图片/视频特征提取核心（Chinese-CLIP 版）
使用 Chinese-CLIP 提取视觉向量 + BLIP 生成描述
支持中文自然语言检索

运行环境要求：
  conda activate aimodel
  Python 3.11.4 | torch 2.0.1+cu117 | GTX 1080 Ti
"""
import json
import torch
import numpy as np
from PIL import Image
import cv2
from pathlib import Path
from typing import Optional

# ── torch 版本守卫（防止在错误的 torch 版本下运行，cu117 专用）──────
_tv = torch.__version__
if not _tv.startswith("2.0") or "cu117" not in _tv:
    raise EnvironmentError(
        f"[Encoder] torch 版本不对！当前: {_tv}\n"
        f"需要 2.0.1+cu117，请运行：\n"
        f"  pip install torch==2.0.1+cu117 torchvision==0.15.2+cu117 "
        f"--index-url https://download.pytorch.org/whl/cu117"
    )

# ── 全局设备 ──────────────────────────────────────────────────
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
print(f"[Encoder] 使用设备: {DEVICE}  (torch {torch.__version__})")

# ── Chinese-CLIP 模型加载 ─────────────────────────────────────
# 可用模型: ViT-B/16(快) ViT-L/14(推荐) ViT-L/14-336(精细) ViT-H/14(最强)
# 中文理解原生支持，不需要中英文映射模板
_CN_CLIP_MODEL = "ViT-L-14"   # ← 如需更高质量改为 "ViT-L-14-336"

_cn_model = None
_cn_preprocess = None

def get_cn_clip_model():
    """加载 Chinese-CLIP 模型（首次运行自动从 hf-mirror 下载）"""
    global _cn_model, _cn_preprocess
    if _cn_model is None:
        # 禁用代理，避免 SOCKS 代理导致 httpx 报错
        os.environ["http_proxy"] = ""
        os.environ["https_proxy"] = ""
        os.environ["HTTP_PROXY"] = ""
        os.environ["HTTPS_PROXY"] = ""
        os.environ.setdefault("HF_ENDPOINT", "https://hf-mirror.com")
        from cn_clip.clip import load_from_name
        print(f"[Encoder] 加载 Chinese-CLIP 模型 ({_CN_CLIP_MODEL})...")
        _cn_model, _cn_preprocess = load_from_name(
            _CN_CLIP_MODEL,
            device=DEVICE,
            download_root=os.path.expanduser("~/.cache/chinese_clip")
        )
        _cn_model.eval()
        # 打印向量维度，用于确认 FAISS 索引维度
        with torch.no_grad():
            dummy = _cn_preprocess(Image.new("RGB", (224, 224))).unsqueeze(0).to(DEVICE)
            feat = _cn_model.encode_image(dummy)
            DIM = feat.shape[-1]
        print(f"[Encoder] Chinese-CLIP 加载完成，向量维度: {DIM}")
        global VEC_DIM
        VEC_DIM = DIM
    return _cn_model, _cn_preprocess


# ── BLIP 图像描述模型（可选，生成图片描述） ────────────────────
_blip_processor = None
_blip_model = None
_BLIP_LOCAL_PATH = None  # 动态获取本地路径

def _get_blip_local_path():
    """获取 BLIP 模型的本地缓存路径（自动下载 if needed）"""
    global _BLIP_LOCAL_PATH
    if _BLIP_LOCAL_PATH is not None:
        return _BLIP_LOCAL_PATH
    from huggingface_hub import snapshot_download
    os.environ.setdefault("HF_ENDPOINT", "https://hf-mirror.com")
    print("[Encoder] 检查 BLIP 模型缓存...")
    cache_dir = snapshot_download(
        repo_id="Salesforce/blip-image-captioning-large",
        cache_dir=os.path.expanduser("~/.cache/huggingface/hub"),
        resume_download=True,
    )
    _BLIP_LOCAL_PATH = cache_dir
    print(f"[Encoder] BLIP 本地路径: {cache_dir}")
    return cache_dir

def get_blip_model():
    """加载 BLIP 模型生成图片描述（英文，可后续翻译）"""
    global _blip_processor, _blip_model
    if _blip_model is None:
        from transformers import BlipProcessor, BlipForConditionalGeneration
        print("[Encoder] 加载 BLIP 描述模型...")
        local_path = _get_blip_local_path()
        _blip_processor = BlipProcessor.from_pretrained(local_path, use_fast=False)
        _blip_model = BlipForConditionalGeneration.from_pretrained(
            local_path, torch_dtype=torch.float16
        ).to(DEVICE)
        _blip_model.eval()
        print("[Encoder] BLIP 加载完成")
    return _blip_processor, _blip_model


# ── 图片编码 ─────────────────────────────────────────────────
VEC_DIM = None  # 在 get_cn_clip_model() 中动态确定

def encode_image(image_path: str) -> Optional[np.ndarray]:
    """提取图片的 Chinese-CLIP 向量"""
    try:
        model, preprocess = get_cn_clip_model()
        img = Image.open(image_path).convert("RGB")
        tensor = preprocess(img).unsqueeze(0).to(DEVICE)
        with torch.no_grad():
            features = model.encode_image(tensor)
            features = features / features.norm(dim=-1, keepdim=True)
        return features.cpu().numpy().astype(np.float32)[0]
    except Exception as e:
        print(f"[Encoder] 图片编码失败 {image_path}: {e}")
        return None


def describe_image(image_path: str) -> str:
    """用 BLIP 生成图片描述"""
    try:
        processor, model = get_blip_model()
        img = Image.open(image_path).convert("RGB")
        img.thumbnail((512, 512), Image.LANCZOS)
        inputs = processor(img, return_tensors="pt").to(DEVICE, torch.float16)
        with torch.no_grad():
            out = model.generate(**inputs, max_new_tokens=60)
        caption = processor.decode(out[0], skip_special_tokens=True)
        return caption
    except Exception as e:
        print(f"[Encoder] 图片描述失败 {image_path}: {e}")
        return ""


# ── 视频处理（抽帧取平均向量） ─────────────────────────────────
def extract_video_frames(video_path: str, max_frames: int = 8) -> list:
    """从视频均匀抽取关键帧，返回 PIL Image 列表"""
    frames = []
    try:
        cap = cv2.VideoCapture(video_path)
        total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        if total <= 0:
            cap.release()
            return frames
        interval = max(1, total // max_frames)
        for i in range(0, min(total, max_frames * interval), interval):
            cap.set(cv2.CAP_PROP_POS_FRAMES, i)
            ret, frame = cap.read()
            if ret:
                rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                frames.append(Image.fromarray(rgb))
            if len(frames) >= max_frames:
                break
        cap.release()
    except Exception as e:
        print(f"[Encoder] 视频抽帧失败 {video_path}: {e}")
    return frames


def encode_video(video_path: str, max_frames: int = 8) -> Optional[np.ndarray]:
    """提取视频的平均 Chinese-CLIP 向量（多帧平均）"""
    frames = extract_video_frames(video_path, max_frames)
    if not frames:
        return None
    model, preprocess = get_cn_clip_model()
    vecs = []
    for img in frames:
        try:
            t = preprocess(img).unsqueeze(0).to(DEVICE)
            with torch.no_grad():
                feat = model.encode_image(t)
                feat = feat / feat.norm(dim=-1, keepdim=True)
            vecs.append(feat.cpu().numpy().astype(np.float32)[0])
        except Exception:
            continue
    if not vecs:
        return None
    return np.mean(vecs, axis=0)


def describe_video(video_path: str, max_frames: int = 4) -> str:
    """对视频抽帧并生成描述（取第一帧）"""
    frames = extract_video_frames(video_path, max_frames)
    if not frames:
        return ""
    try:
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as f:
            tmp = f.name
        frames[0].save(tmp)
        caption = describe_image(tmp)
        os.unlink(tmp)
        return caption
    except Exception:
        return ""


# ── 文本编码（用于检索，原生支持中文） ─────────────────────────
def encode_text(text: str) -> np.ndarray:
    """将文本（中英文均可）编码为向量，用于语义检索"""
    from cn_clip.clip import tokenize
    model, _ = get_cn_clip_model()
    tokens = tokenize([text]).to(DEVICE)
    with torch.no_grad():
        feat = model.encode_text(tokens)
        feat = feat / feat.norm(dim=-1, keepdim=True)
    return feat.cpu().numpy().astype(np.float32)[0]


# ── 批量编码（建索引时加速） ──────────────────────────────────
def encode_images_batch(image_paths: list, batch_size: int = 16) -> list:
    """批量编码图片，返回向量列表（断点续传友好）"""
    model, preprocess = get_cn_clip_model()
    results = []
    for i in range(0, len(image_paths), batch_size):
        batch = image_paths[i:i+batch_size]
        tensors = []
        valid = []
        for p in batch:
            try:
                img = Image.open(p).convert("RGB")
                tensors.append(preprocess(img))
                valid.append(p)
            except Exception:
                results.append(None)
        if not tensors:
            continue
        batch_tensor = torch.stack(tensors).to(DEVICE)
        with torch.no_grad():
            feats = model.encode_image(batch_tensor)
            feats = feats / feats.norm(dim=-1, keepdim=True)
        for p, f in zip(valid, feats):
            results.append(f.cpu().numpy().astype(np.float32)[0])
            # 补齐 skipped
            while len(results) < image_paths.index(p) + 1:
                results.insert(image_paths.index(p), None)
    return results


# ══════════════════════════════════════════════════════════════════
#  功能1: EasyOCR - 扫描型 PDF / 图片文字识别
#  （PaddleOCR 依赖过于复杂，EasyOCR 更稳定，支持中文+GPU）
# ══════════════════════════════════════════════════════════════════
_EASY_OCR_ENGINE = None

def get_easy_ocr():
    """懒加载 EasyOCR（首次使用时初始化，GPU 加速）"""
    global _EASY_OCR_ENGINE
    if _EASY_OCR_ENGINE is None:
        try:
            import easyocr
            print("[Encoder] 初始化 EasyOCR（中文+英文，GPU 加速）...")
            _EASY_OCR_ENGINE = easyocr.Reader(
                ['ch_sim', 'en'],      # 简体中文 + 英文
                gpu=True,              # GPU 加速（GTX 1080 Ti）
                model_storage_directory=os.path.expanduser("~/.cache/easyocr"),
                download_enabled=True,
            )
            print("[Encoder] EasyOCR 初始化完成")
        except Exception as e:
            print(f"[Encoder] EasyOCR 初始化失败: {e}")
            _EASY_OCR_ENGINE = None
    return _EASY_OCR_ENGINE

def ocr_image(image_path: str) -> str:
    """用 EasyOCR 识别图片中的文字"""
    try:
        engine = get_easy_ocr()
        if engine is None:
            return ""
        result = engine.readtext(image_path)
        if not result:
            return ""
        lines = []
        for detection in result:
            if len(detection) >= 2:
                text = detection[1].strip()
                if text:
                    lines.append(text)
        return "\n".join(lines)
    except Exception as e:
        print(f"[Encoder] OCR 识别失败 {image_path}: {e}")
        return ""

def ocr_pdf_pages(pdf_path: str, max_pages: int = 20) -> str:
    """将 PDF 页面转为图片，用 EasyOCR 识别文字（扫描型 PDF）"""
    try:
        from pdf2image import convert_from_path
        engine = get_easy_ocr()
        if engine is None:
            return ""
        images = convert_from_path(pdf_path, dpi=150, first_page=1, last_page=max_pages)
        all_text = []
        for i, img in enumerate(images):
            try:
                import tempfile, os
                with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False, dir=os.environ.get("TEMP", ".")) as f:
                    tmp = f.name
                img.save(tmp, "JPEG", quality=85)
                text = ocr_image(tmp)
                if text:
                    all_text.append(text)
                os.unlink(tmp)
            except Exception as e:
                print(f"[Encoder] PDF 第{i+1}页 OCR 失败: {e}")
        return "\n".join(all_text)
    except Exception as e:
        print(f"[Encoder] PDF OCR 失败 {pdf_path}: {e}")
        return ""


# ══════════════════════════════════════════════════════════════════
#  功能2: 文档分块 - 长文档按段落/字符数切分
# ══════════════════════════════════════════════════════════════════
def chunk_text(text: str, chunk_size: int = 300, overlap: int = 50) -> list:
    """
    将长文本切分为多个块（按段落边界或固定字符数）
    
    Args:
        text: 原始文本
        chunk_size: 每块最大字符数
        overlap: 相邻块重叠字符数
    
    Returns:
        文本块列表，每块含文本内容和位置信息
    """
    if not text or len(text) <= chunk_size:
        return [{"text": text.strip(), "start": 0, "chunk_id": 0}]
    
    chunks = []
    # 先按段落分割
    paragraphs = text.split("\n")
    current_chunk = ""
    chunk_id = 0
    
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        # 如果单个段落就超长，按句子继续切
        if len(para) > chunk_size:
            # 先保存当前累积的块
            if current_chunk.strip():
                chunks.append({
                    "text": current_chunk.strip(),
                    "start": 0,
                    "chunk_id": chunk_id
                })
                chunk_id += 1
                current_chunk = ""
            # 按句子切分超长段落
            for i in range(0, len(para), chunk_size - overlap):
                sub = para[i:i + chunk_size]
                chunks.append({
                    "text": sub.strip(),
                    "start": i,
                    "chunk_id": chunk_id
                })
                chunk_id += 1
        elif len(current_chunk) + len(para) + 1 <= chunk_size:
            current_chunk += ("\n" if current_chunk else "") + para
        else:
            # 当前块已满，保存并开新块
            if current_chunk.strip():
                chunks.append({
                    "text": current_chunk.strip(),
                    "start": 0,
                    "chunk_id": chunk_id
                })
                chunk_id += 1
            current_chunk = para
    
    # 保存最后一块
    if current_chunk.strip():
        chunks.append({
            "text": current_chunk.strip(),
            "start": 0,
            "chunk_id": chunk_id
        })
    
    return chunks if chunks else [{"text": text.strip()[:chunk_size], "start": 0, "chunk_id": 0}]


# ══════════════════════════════════════════════════════════════════
#  文档文本提取（支持扫描型 PDF）
# ══════════════════════════════════════════════════════════════════

def extract_pdf_text(file_path: str, max_pages: int = 50, use_ocr_fallback: bool = True) -> str:
    """提取 PDF 文本内容（支持文本型 PDF，文本为空时自动回退到 OCR）"""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= max_pages:
                    break
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        text = "\n".join(text_parts).strip()

        # 功能1: 如果文本为空（扫描型PDF），自动回退到 OCR（EasyOCR）
        if (not text or len(text) < 50) and use_ocr_fallback:
            print(f"[Encoder] PDF 文本内容过少，启用 EasyOCR: {file_path}")
            text = ocr_pdf_pages(file_path, max_pages=max_pages)

        return text
    except Exception as e:
        print(f"[Encoder] PDF文本提取失败 {file_path}: {e}")
        # 回退到 OCR（EasyOCR）
        if use_ocr_fallback:
            return ocr_pdf_pages(file_path, max_pages=max_pages)
        return ""

def extract_pdf_images(file_path: str, max_pages: int = 5) -> list:
    """提取 PDF 页面为图像（用于 OCR 扫描版 PDF）"""
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(file_path, dpi=150, first_page=1, last_page=max_pages)
        return images
    except Exception as e:
        print(f"[Encoder] PDF转图像失败 {file_path}: {e}")
        return []

def extract_docx_text(file_path: str) -> str:
    """提取 Word 文档文本内容"""
    try:
        from docx import Document
        doc = Document(file_path)
        text_parts = []
        for para in doc.paragraphs:
            if para.text.strip():
                text_parts.append(para.text.strip())
        return "\n".join(text_parts).strip()
    except Exception as e:
        print(f"[Encoder] Word文本提取失败 {file_path}: {e}")
        return ""

def extract_txt_text(file_path: str) -> str:
    """提取 TXT 文件内容"""
    try:
        # 尝试多种编码
        for enc in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    return f.read(50000).strip()  # 最多读5万字
            except UnicodeDecodeError:
                continue
        return ""
    except Exception as e:
        print(f"[Encoder] TXT文本提取失败 {file_path}: {e}")
        return ""

def extract_xlsx_text(file_path: str, max_rows: int = 200) -> str:
    """提取 Excel 文件文本内容"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames[:3]:  # 最多读3个工作表
            ws = wb[sheet_name]
            row_count = 0
            for row in ws.iter_rows(values_only=True, max_row=max_rows):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    text_parts.append(" | ".join(cells))
                row_count += 1
                if row_count >= max_rows:
                    break
        wb.close()
        return "\n".join(text_parts).strip()
    except Exception as e:
        print(f"[Encoder] Excel文本提取失败 {file_path}: {e}")
        return ""

def extract_document_text(file_path: str) -> str:
    """通用文档文本提取入口，根据扩展名调用对应函数"""
    ext = Path(file_path).suffix.lower()
    if ext == '.pdf':
        return extract_pdf_text(file_path)
    elif ext in ['.docx', '.doc']:
        return extract_docx_text(file_path)
    elif ext == '.txt':
        return extract_txt_text(file_path)
    elif ext in ['.xlsx', '.xls']:
        return extract_xlsx_text(file_path)
    else:
        return ""

def encode_document(file_path: str, chunk_size: int = 300) -> list:
    """
    提取文档文本，分块，每块编码为向量（功能2：文档分块索引）
    
    Returns:
        list of dicts: [{"chunk_text": str, "chunk_id": int, "vector": np.ndarray}, ...]
        如果文本太短无法分块，返回空列表
    """
    try:
        text = extract_document_text(file_path)
        if not text or len(text) < 10:
            return []
        
        # 功能2: 按段落/字符数分块
        chunks = chunk_text(text, chunk_size=chunk_size)
        
        results = []
        for chunk in chunks:
            chunk_text_str = chunk["text"]
            if len(chunk_text_str) < 5:  # 跳过太短的块
                continue
            # 用 Chinese-CLIP 文本编码器编码（前512字符）
            vec = encode_text(chunk_text_str[:512])
            results.append({
                "chunk_text": chunk_text_str,
                "chunk_id": chunk["chunk_id"],
                "vector": vec,
            })
        return results
    except Exception as e:
        print(f"[Encoder] 文档编码失败 {file_path}: {e}")
        return []


def encode_document_single(file_path: str) -> Optional[np.ndarray]:
    """
    提取文档文本并编码为单个向量（兼容旧接口）
    仅用于文件级别索引，不再用于分块文档
    """
    try:
        text = extract_document_text(file_path)
        if not text or len(text) < 10:
            return None
        vec = encode_text(text[:512])
        return vec
    except Exception as e:
        print(f"[Encoder] 文档编码失败 {file_path}: {e}")
        return None


if __name__ == "__main__":
    # 快速测试（使用英文避免 Windows 控制台编码问题）
    print("\n── Test text encoding (Chinese) ──")
    vec_cn = encode_text("一只猫坐在沙发上")
    print(f"Chinese vector shape: {vec_cn.shape}")

    print("\n── Test text encoding (English) ──")
    vec_en = encode_text("a cat sitting on a sofa")
    print(f"English vector shape: {vec_en.shape}")

    sim = np.dot(vec_cn, vec_en)
    print(f"Zh/En semantic similarity: {sim:.4f}")
    print("\n[OK] Chinese-CLIP test passed!")
